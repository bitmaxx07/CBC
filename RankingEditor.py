import openpyxl
from tkinter import *
from PIL import Image
from tkinter import messagebox, filedialog

# 2-image, 3-name, 4-match, 5-win, 6-draw, 7-lose, 8-goal, 9-against, 10-difference, 11-score

file = ""


def set_file(filename):
    file = filename


class Team(object):
    def __init__(self, image, name, match, win, draw, lose, goal, against, difference, score):
        self.image = image
        self.name = name
        self.match = match
        self.win = win
        self.draw = draw
        self.lose = lose
        self.goal = goal
        self.against = against
        self.difference = difference
        self.score = score

    def print_all_info(self):
        print("image: " + self.image)
        print("name: " + self.name)
        print("win: " + str(self.win))
        print("draw: " + str(self.draw))
        print("lose: " + str(self.lose))
        print("goal: " + str(self.goal))
        print("against: " + str(self.against))
        print("difference: " + str(self.difference))
        print("score: " + str(self.score))


def evaluate(a, score_a,  b, score_b):
    a.match += 1
    b.match += 1
    a.goal = a.goal + score_a
    a.against = a.against + score_b
    b.goal = b.goal + score_b
    b.against = b.against + score_a
    a.difference = a.goal - a.against
    b.difference = b.goal - b.against

    if score_a > score_b:
        a.win = a.win + 1
        b.lose = b.lose + 1
        a.score = a.score + 3

    elif score_a < score_b:
        b.win = b.win + 1
        a.lose = a.lose + 1
        b.score = b.score + 3

    else:
        a.draw = a.draw + 1
        b.draw = b.draw + 1
        a.score = a.score + 1
        b.score = b.score + 1


def fill_in_sheet(team, group, rank):
    if group == "A":
        ws = ws_a
    elif group == "B":
        ws = ws_b
    else:
        ws = ws_c

    resize_img = Image.open(team.image)
    new_image = resize_img.resize((80, 80))
    new_image.save(team.image.split(".png")[0] + "_new.png")

    img = openpyxl.drawing.image.Image(team.image.split(".png")[0] + "_new.png")
    img.anchor = "B" + str(rank + 2)
    ws.add_image(img)

    ws.cell(rank + 2, 3).value = team.name
    ws.cell(rank + 2, 4).value = team.match
    ws.cell(rank + 2, 5).value = team.win
    ws.cell(rank + 2, 6).value = team.draw
    ws.cell(rank + 2, 7).value = team.lose
    ws.cell(rank + 2, 8).value = team.goal
    ws.cell(rank + 2, 9).value = team.against
    ws.cell(rank + 2, 10).value = team.difference
    ws.cell(rank + 2, 11).value = team.score


def ranking(team_a, team_b, team3, team4):
    ranks = [team_a, team_b, team3, team4]
    ranks = sorted(ranks, key=lambda x: (x.score, x.goal, x.win), reverse=True)
    return ranks


team_dic = {"D": "多特蒙德CFD 13华人足球队", "BO": "波鸿原点Ppagei华人足球队",
            "N": "KSC弗兰肯足球联队", "DU": "打酱油杜伊斯堡队",
            "B": "柏林华人足球队", "S": "斯图加特华人足球队",
            "CH": "开姆尼茨华人足球队", "L": "卢森堡华人足球协会",
            "M": "慕尼黑华人联合足球俱乐部", "F": "法兰克福坚强足球队",
            "DD": "德累斯顿CFC华人足球队", "SCH": "Schöneberg华人足球队"}

image_dic = {"D": "正方形/多特蒙德.png", "BO": "正方形/波鸿.png",
            "N": "正方形/纽伦堡.png", "DU": "正方形/杜伊斯堡.png",
            "B": "正方形/柏林.png", "S": "正方形/斯图加特.png",
            "CH": "正方形/开姆尼茨.png", "L": "正方形/卢森堡.png",
            "M": "正方形/慕尼黑.png", "F": "正方形/法兰克福.png",
            "DD": "正方形/德累斯顿.png", "SCH": "正方形/Schoeneberg.png"}

wb = openpyxl.load_workbook("小组积分榜.xlsx")
ws_a = wb["A小组"]
ws_b = wb["B小组"]
ws_c = wb["C小组"]

team_D = Team(image_dic["D"], team_dic["D"], 0, 0, 0, 0, 0, 0, 0, 0)
team_BO = Team(image_dic["BO"], team_dic["BO"], 0, 0, 0, 0, 0, 0, 0, 0)
team_N = Team(image_dic["N"], team_dic["N"], 0, 0, 0, 0, 0, 0, 0, 0)
team_DU = Team(image_dic["DU"], team_dic["DU"], 0, 0, 0, 0, 0, 0, 0, 0)
team_B = Team(image_dic["B"], team_dic["B"], 0, 0, 0, 0, 0, 0, 0, 0)
team_S = Team(image_dic["S"], team_dic["S"], 0, 0, 0, 0, 0, 0, 0, 0)
team_CH = Team(image_dic["CH"], team_dic["CH"], 0, 0, 0, 0, 0, 0, 0, 0)
team_L = Team(image_dic["L"], team_dic["L"], 0, 0, 0, 0, 0, 0, 0, 0)
team_M = Team(image_dic["M"], team_dic["M"], 0, 0, 0, 0, 0, 0, 0, 0)
team_F = Team(image_dic["F"], team_dic["F"], 0, 0, 0, 0, 0, 0, 0, 0)
team_DD = Team(image_dic["DD"], team_dic["DD"], 0, 0, 0, 0, 0, 0, 0, 0)
team_SCH = Team(image_dic["SCH"], team_dic["SCH"], 0, 0, 0, 0, 0, 0, 0, 0)

total_team_list = [team_D, team_BO, team_N, team_DU, team_B, team_S, team_CH, team_L, team_M, team_F, team_DD, team_SCH]


class Window(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.var = StringVar(master)

        self.menu = OptionMenu(master, self.var, "A", "B", "C", command=self.callback)
        self.var.set("A")
        self.temp = "A"
        print(self.temp)
        self.menu.pack()
        # self.vcmd = master.register(self.callback)
        self.team_list = []
        self.var_team1 = StringVar(master)
        self.var_team2 = StringVar(master)

        self.team_A = OptionMenu(master, self.var_team1, "")
        self.team_A.place(x=40, y=60)
        self.score_A = Entry(master, width=10)
        self.score_A.place(x=100, y=60)
        self.team_B = OptionMenu(master, self.var_team2, "")
        self.team_B.place(x=40, y=100)
        self.score_B = Entry(master, width=10)
        self.score_B.place(x=100, y=100)
        self.sc_a = self.score_A.get()
        self.sc_b = self.score_B.get()
        self.button = Button(master, text="start", command=self.start).place(x=100, y=150)
        self.refresh_button = Button(master, text="refresh", command=self.refresh).place(x=25, y=20)
        # self.file = ""
        self.select_file_button = Button(master, text="select file", command=self.load_xlsx).place(x=100, y=20)
        self.pack()

    def callback(self, selection):
        self.var.set(selection)
        self.temp = selection
        # print(self.temp)

    def load_xlsx(self):
        self.file = filedialog.askopenfilename(title="Choose excel file", filetypes=[("Excel", "*.xlsx")])

    def refresh(self):
        self.team_A = OptionMenu(self.master, self.var_team1, *self.get_var())
        self.team_A.place(x=40, y=60)
        self.team_B = OptionMenu(self.master, self.var_team2, *self.get_var())
        self.team_B.place(x=40, y=100)
        print(self.get_var())

    def get_var(self):
        if self.temp == "A":
            self.team_list = ["D", "BO", "N", "DU"]
            return self.team_list
        elif self.temp == "B":
            self.team_list = ["B", "S", "CH", "L"]
            return self.team_list
        else:
            self.team_list = ["M", "F", "DD", "SCH"]
            return self.team_list

    def start(self):
        global team1, team2
        self.sc_a = self.score_A.get()
        self.sc_b = self.score_B.get()
        if str.isdigit(self.sc_a) is not True or str.isdigit(self.sc_b) is not True:
            messagebox.showinfo("Score must be a digit!")
        else:
            wb = openpyxl.load_workbook(self.file.name)
            temp_team_list = []
            if self.var_team1.get() == self.var_team2.get():
                messagebox.showwarning("cannot deal with two same teams!")
            for team in total_team_list:
                if team_dic[self.var_team1.get()] == team.name:
                    team1 = team
                if team_dic[self.var_team2.get()] == team.name:
                    team2 = team
                for t in self.team_list:
                    if team_dic[self.var_team1.get()] != team.name and team_dic[self.var_team2.get()] != team.name:
                        if team_dic[t] == team.name:
                            temp_team_list.append(team)
            evaluate(team1, int(self.sc_a), team2, int(self.sc_b))
            '''temp_list = self.team_list
            temp_list.remove(self.var_team1.get())
            temp_list.remove(self.var_team2.get())'''
            rank_list = ranking(team1, team2, temp_team_list[0], temp_team_list[1])
            for ts in rank_list:
                fill_in_sheet(ts, self.temp, rank_list.index(ts) + 1)

            wb.save("result.xlsx")
            messagebox.showinfo("Done!")

    '''def callback(self, p):
        if str.isdigit(p) or p == "":
            return True
        else:
            return False'''


'''team_a = Team(image_dic["D"], team_dic["D"], 0, 0, 0, 0, 0, 0, 0, 0)
team_b = Team(image_dic["N"], team_dic["N"], 0, 0, 0, 0, 0, 0, 0, 0)
team_c = Team(image_dic["B"], team_dic["B"], 0, 0, 0, 0, 0, 0, 0, 0)
team_d = Team(image_dic["DU"], team_dic["DU"], 0, 0, 0, 0, 0, 0, 0, 0)

evaluate(team_a, 2, team_b, 0)
evaluate(team_c, 3, team_d, 1)

team_a.print_all_info()
print()
team_b.print_all_info()'''

'''fill_in_sheet(team_a, "A", 2)
fill_in_sheet(team_b, "A", 4)
fill_in_sheet(team_c, "A", 1)
fill_in_sheet(team_d, "A", 3)'''
'''rankings = ranking(team_a, team_b, team_c, team_d)

for t in rankings:
    t.print_all_info()
    fill_in_sheet(t, "A", rankings.index(t) + 1)'''

# wb.save("result.xlsx")

root = Tk()
Window(root)
root.geometry("300x250")
root.mainloop()

